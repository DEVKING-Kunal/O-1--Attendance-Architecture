"""
sync.py — Post-session cloud sync.

Changes from v1:
  - Each session now gets its OWN worksheet tab inside the master
    "Attendance_Sheet" spreadsheet (named after the session timestamp),
    instead of every session being appended into one growing sheet1.
    Professors can open any past session independently.
  - Every session is ALSO exported as a standalone, formatted .xlsx file
    in output/xlsx/ — so a professor can grab attendance immediately,
    even with no internet connection and without ever opening Sheets.
    The cloud push is now a bonus on top of that, not a dependency.
  - The local .xlsx is written BEFORE the cloud call, so a failed/slow
    internet connection never blocks a professor from having the file.
"""

import csv
import os
import sys
import time
from pathlib import Path

import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path("output")
XLSX_DIR = OUTPUT_DIR / "xlsx"
SPREADSHEET_NAME = "Attendance_Sheet"   # the parent Google Sheet; tabs are added per session


def get_latest_attendance_file():
    candidates = list(OUTPUT_DIR.glob("attendance_*.csv"))
    if not candidates:
        # backward compatibility with v1, which wrote CSVs to the project root
        candidates = [Path(f) for f in os.listdir(".") if f.startswith("attendance_") and f.endswith(".csv")]
    return max(candidates, key=lambda p: p.stat().st_ctime) if candidates else None


def session_name_from_filename(csv_path: Path) -> str:
    # attendance_2026-07-02_10-15.csv -> 2026-07-02_10-15
    return csv_path.stem.replace("attendance_", "", 1)


def read_rows(csv_path: Path):
    with open(csv_path, mode="r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader)
        rows = list(reader)
    return header, rows


def export_xlsx(session_name: str, header, rows) -> Path:
    XLSX_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = session_name[:31]  # Excel sheet-name limit is 31 chars

    ws.append(header)
    header_fill = PatternFill("solid", fgColor="1A73E8")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append(row)

    for i, col_name in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(col_name) + 2)

    xlsx_path = XLSX_DIR / f"{session_name}.xlsx"
    wb.save(xlsx_path)
    return xlsx_path


def push_to_sheets(session_name: str, header, rows) -> str:
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = ServiceAccountCredentials.from_json_keyfile_name("credentials.json", scope)
    client = gspread.authorize(creds)
    spreadsheet = client.open(SPREADSHEET_NAME)

    tab_name = session_name[:100]  # Google Sheets tab-name limit
    try:
        worksheet = spreadsheet.worksheet(tab_name)
        # Tab already exists — likely a retry after a previous partial failure.
    except gspread.WorksheetNotFound:
        worksheet = spreadsheet.add_worksheet(
            title=tab_name, rows=str(len(rows) + 10), cols=str(len(header) + 2)
        )

    worksheet.append_row(header)
    if rows:
        worksheet.append_rows(rows)
    return worksheet.url


def sync_to_cloud():
    csv_path = get_latest_attendance_file()
    if not csv_path:
        print("No local attendance files found.")
        return

    session_name = session_name_from_filename(csv_path)
    print(f"Starting sync for session: {session_name}  ({csv_path})")

    header, rows = read_rows(csv_path)
    if not rows:
        print("File is empty. Skipping.")
        return

    # 1. Local .xlsx first — succeeds even fully offline.
    xlsx_path = export_xlsx(session_name, header, rows)
    print(f"📄 Local Excel copy ready: {xlsx_path}")

    # 2. Cloud push is best-effort on top of that.
    try:
        sheet_url = push_to_sheets(session_name, header, rows)
        print(f"✅ Synced {len(rows)} records to a new tab in Google Sheets: {sheet_url}")

        backup_name = OUTPUT_DIR / f"synced_backup_{int(time.time())}_{csv_path.name}"
        csv_path.rename(backup_name)
        print(f"CSV archived as: {backup_name}")
    except Exception as e:
        print(f"❌ CLOUD SYNC FAILED: {e}")
        print("⚠️ Local CSV preserved for retry. The .xlsx copy already exists and is unaffected.")


if __name__ == "__main__":
    sync_to_cloud()